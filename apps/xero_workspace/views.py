import ast
from datetime import datetime
from zipfile import BadZipFile

import openpyxl
import psycopg2
from django.contrib import messages
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import IntegrityError
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.utils.timezone import make_aware
from django.views import View
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.task_log.models import TaskLog
from apps.xero_workspace.forms import CategoryMappingForm, EmployeeMappingForm, ScheduleForm, ProjectMappingForm
from apps.xero_workspace.models import Workspace, CategoryMapping, EmployeeMapping, \
    WorkspaceSchedule, ProjectMapping
from apps.xero_workspace.tasks import schedule_sync, run_sync_schedule


class WorkspaceView(View):
    """
    Workspace Dashboard view
    """
    template_name = "xero_workspace/workspace.html"

    def dispatch(self, request, *args, **kwargs):
        method = self.request.POST.get('method', '').lower()
        if method == 'delete':
            return self.delete(request)
        return super(WorkspaceView, self).dispatch(request, *args, **kwargs)

    def get(self, request):
        user_workspaces = Workspace.objects.filter(user=request.user).order_by('-created_at')
        for workspace in user_workspaces:
            task_log = TaskLog.objects.filter(workspace=workspace)
            if task_log:
                workspace.last_sync = task_log.latest().created_at
            else:
                workspace.last_sync = "-"
        page = request.GET.get('page', 1)
        paginator = Paginator(user_workspaces, 10)
        try:
            user_workspaces = paginator.page(page)
        except PageNotAnInteger:
            user_workspaces = paginator.page(1)
        except EmptyPage:
            user_workspaces = paginator.page(paginator.num_pages)
        return render(request, self.template_name, {"workspaces": user_workspaces})

    def post(self, request):
        workspace_name = request.POST.get('new-workspace-name')
        workspace = Workspace.objects.create(name=workspace_name)
        workspace.user.add(request.user)
        workspace.save()
        return HttpResponseRedirect(self.request.path_info)

    def delete(self, request):
        selected_workspace = [ast.literal_eval(x) for x in request.POST.getlist('workspace_ids')]
        Workspace.objects.filter(id__in=selected_workspace).delete()
        return HttpResponseRedirect(self.request.path_info)


class CategoryMappingView(View):
    """
    Category Mapping View
    """
    template_name = "xero_workspace/category_mapping.html"
    workspace = None

    def dispatch(self, request, *args, **kwargs):
        method = self.request.POST.get('method', '').lower()
        if method == 'update':
            return self.update(request, *args, **kwargs)
        if method == 'delete':
            return self.delete(request, *args, **kwargs)
        return super(CategoryMappingView, self).dispatch(request, *args, **kwargs)

    def setup(self, request, *args, **kwargs):
        workspace_id = kwargs['workspace_id']
        self.workspace = Workspace.objects.get(id=workspace_id)
        super(CategoryMappingView, self).setup(request)

    def get(self, request, workspace_id):
        category_mappings = CategoryMapping.objects.filter(workspace__id=workspace_id).order_by('-created_at')
        page = request.GET.get('page', 1)
        paginator = Paginator(category_mappings, 10)
        try:
            category_mappings = paginator.page(page)
        except PageNotAnInteger:
            category_mappings = paginator.page(1)
        except EmptyPage:
            category_mappings = paginator.page(paginator.num_pages)
        form = CategoryMappingForm()
        context = {"category_mapping": "active", "form": form,
                   "mappings": category_mappings, "mappings_tab": "active"}
        return render(request, self.template_name, context)

    def post(self, request, workspace_id):
        form = CategoryMappingForm(request.POST)
        if form.is_valid:
            category = request.POST.get('category')
            sub_category = request.POST.get('sub_category')
            account_code = request.POST.get('account_code')
            try:
                CategoryMapping.objects.create(workspace=self.workspace,
                                               category=category,
                                               sub_category=sub_category,
                                               account_code=account_code)
            except (psycopg2.errors.UniqueViolation, IntegrityError):
                messages.error(request, 'Mapping already exists')
        return HttpResponseRedirect(self.request.path_info)

    def update(self, request, workspace_id):
        form = CategoryMappingForm(request.POST)
        if form.is_valid:
            category = request.POST.get('category')
            sub_category = request.POST.get('sub_category')
            account_code = request.POST.get('account_code')
            mapping_id = request.POST.get('mapping_id')
            CategoryMapping.objects.filter(id=mapping_id).update(category=category,
                                                                 sub_category=sub_category,
                                                                 account_code=account_code,
                                                                 invalid=False)
        return HttpResponseRedirect(self.request.path_info)

    def delete(self, request, workspace_id):
        selected_mappings = [ast.literal_eval(x) for x in request.POST.getlist('mapping_ids')]
        CategoryMapping.objects.filter(id__in=selected_mappings).delete()
        return HttpResponseRedirect(self.request.path_info)


class CategoryMappingBulkUploadView(View):
    """
    Category mapping bulk upload view
    """

    @staticmethod
    def post(request, workspace_id):
        workspace = Workspace.objects.get(id=workspace_id)
        file = request.FILES['bulk_upload_file']
        try:
            work_book = openpyxl.load_workbook(file)
            worksheet = work_book.active
            for category, sub_category, account_code in worksheet.iter_rows(min_row=2):
                sub_category.value = '' if sub_category.value is None else sub_category.value
                CategoryMapping.objects.update_or_create(workspace=workspace, category=category.value,
                                                         sub_category=sub_category.value,
                                                         defaults={'account_code': account_code.value})
        except (ValueError, BadZipFile, KeyError):
            messages.error(request, 'The uploaded file has invalid column(s): Please upload again')
        return HttpResponseRedirect(reverse('xero_workspace:category_mapping', args=[workspace_id]))


class EmployeeMappingView(View):
    """
    Employee Mapping View
    """
    template_name = "xero_workspace/employee_mapping.html"
    workspace = None

    def setup(self, request, *args, **kwargs):
        workspace_id = kwargs['workspace_id']
        self.workspace = Workspace.objects.get(id=workspace_id)
        super(EmployeeMappingView, self).setup(request)

    def dispatch(self, request, *args, **kwargs):
        method = self.request.POST.get('method', '').lower()
        if method == 'update':
            return self.update(request, *args, **kwargs)
        if method == 'delete':
            return self.delete(request, *args, **kwargs)
        return super(EmployeeMappingView, self).dispatch(request, *args, **kwargs)

    def get(self, request, workspace_id):
        employee_mappings = EmployeeMapping.objects.filter(workspace__id=workspace_id).order_by('-created_at')
        page = request.GET.get('page', 1)
        paginator = Paginator(employee_mappings, 10)
        try:
            employee_mappings = paginator.page(page)
        except PageNotAnInteger:
            employee_mappings = paginator.page(1)
        except EmptyPage:
            employee_mappings = paginator.page(paginator.num_pages)
        form = EmployeeMappingForm()
        context = {"employee_mapping": "active", "form": form,
                   "mappings": employee_mappings, "mappings_tab": "active"}
        return render(request, self.template_name, context)

    def post(self, request, workspace_id):
        form = EmployeeMappingForm(request.POST)
        if form.is_valid:
            employee_email = request.POST.get('employee_email')
            contact_name = request.POST.get('contact_name')
            try:
                EmployeeMapping.objects.create(workspace=self.workspace,
                                               employee_email=employee_email,
                                               contact_name=contact_name)
            except (psycopg2.errors.UniqueViolation, IntegrityError):
                messages.error(request, 'Mapping already exists')
        return HttpResponseRedirect(self.request.path_info)

    def update(self, request, workspace_id):
        form = EmployeeMappingForm(request.POST)
        if form.is_valid:
            employee_email = request.POST.get('employee_email')
            contact_name = request.POST.get('contact_name')
            mapping_id = request.POST.get('mapping_id')
            EmployeeMapping.objects.filter(id=mapping_id).update(
                employee_email=employee_email, contact_name=contact_name, invalid=False)
        return HttpResponseRedirect(self.request.path_info)

    def delete(self, request, workspace_id):
        selected_mappings = [ast.literal_eval(x) for x in request.POST.getlist('mapping_ids')]
        EmployeeMapping.objects.filter(id__in=selected_mappings).delete()
        return HttpResponseRedirect(self.request.path_info)


class EmployeeMappingBulkUploadView(View):
    """
    Employee mapping bulk upload view
    """

    @staticmethod
    def post(request, workspace_id):
        workspace = Workspace.objects.get(id=workspace_id)
        file = request.FILES['bulk_upload_file']
        try:
            work_book = openpyxl.load_workbook(file)
            worksheet = work_book.active
            for employee_email, contact_name in worksheet.iter_rows(min_row=2):
                EmployeeMapping.objects.update_or_create(workspace=workspace, employee_email=employee_email.value,
                                                         defaults={'contact_name': contact_name.value})
        except(ValueError, BadZipFile, KeyError):
            messages.error(request, 'The uploaded file has invalid column(s): Please upload again')
        return HttpResponseRedirect(reverse('xero_workspace:employee_mapping', args=[workspace_id]))


class ProjectMappingView(View):
    """
    Project Mapping View
    """
    template_name = "xero_workspace/project_mapping.html"
    workspace = None

    def dispatch(self, request, *args, **kwargs):
        method = self.request.POST.get('method', '').lower()
        if method == 'update':
            return self.update(request, *args, **kwargs)
        if method == 'delete':
            return self.delete(request, *args, **kwargs)
        return super(ProjectMappingView, self).dispatch(request, *args, **kwargs)

    def setup(self, request, *args, **kwargs):
        workspace_id = kwargs['workspace_id']
        self.workspace = Workspace.objects.get(id=workspace_id)
        super(ProjectMappingView, self).setup(request)

    def get(self, request, workspace_id):
        project_mappings = ProjectMapping.objects.filter(workspace__id=workspace_id).order_by('-created_at')
        page = request.GET.get('page', 1)
        paginator = Paginator(project_mappings, 10)
        try:
            project_mappings = paginator.page(page)
        except PageNotAnInteger:
            project_mappings = paginator.page(1)
        except EmptyPage:
            project_mappings = paginator.page(paginator.num_pages)
        form = ProjectMappingForm()
        context = {"project_mapping": "active", "form": form,
                   "mappings": project_mappings, "mappings_tab": "active"}
        return render(request, self.template_name, context)

    def post(self, request, workspace_id):
        form = ProjectMappingForm(request.POST)
        if form.is_valid:
            project_name = request.POST.get('project_name')
            tracking_category_name = request.POST.get('tracking_category_name')
            tracking_category_option = request.POST.get('tracking_category_option')
            try:
                ProjectMapping.objects.create(workspace=self.workspace,
                                              project_name=project_name,
                                              tracking_category_name=tracking_category_name,
                                              tracking_category_option=tracking_category_option)
            except (psycopg2.errors.UniqueViolation, IntegrityError):
                messages.error(request, 'Mapping already exists')
        return HttpResponseRedirect(self.request.path_info)

    def update(self, request, workspace_id):
        form = ProjectMappingForm(request.POST)
        if form.is_valid:
            project_name = request.POST.get('project_name')
            tracking_category_name = request.POST.get('tracking_category_name')
            tracking_category_option = request.POST.get('tracking_category_option')
            mapping_id = request.POST.get('mapping_id')
            ProjectMapping.objects.filter(id=mapping_id).update(project_name=project_name,
                                                                tracking_category_name=tracking_category_name,
                                                                tracking_category_option=tracking_category_option,
                                                                invalid=False)

        return HttpResponseRedirect(self.request.path_info)

    def delete(self, request, workspace_id):
        selected_mappings = [ast.literal_eval(x) for x in request.POST.getlist('mapping_ids')]
        ProjectMapping.objects.filter(id__in=selected_mappings).delete()
        return HttpResponseRedirect(self.request.path_info)


class ProjectMappingBulkUploadView(View):
    """
    Project mapping bulk upload view
    """

    @staticmethod
    def post(request, workspace_id):
        workspace = Workspace.objects.get(id=workspace_id)
        file = request.FILES['bulk_upload_file']
        try:
            work_book = openpyxl.load_workbook(file)
            worksheet = work_book.active
            for project_name, tracking_category_name, tracking_category_option in worksheet.iter_rows(min_row=2):
                ProjectMapping.objects.update_or_create(workspace=workspace, project_name=project_name.value, defaults={
                    'tracking_category_name': tracking_category_name.value,
                    'tracking_category_option': tracking_category_option.value})
        except (ValueError, BadZipFile, KeyError):
            messages.error(request, 'The uploaded file has invalid column(s): Please upload again')
        return HttpResponseRedirect(reverse('xero_workspace:project_mapping', args=[workspace_id]))


class ScheduleView(View):
    """
    Schedule View
    """
    template_name = "xero_workspace/schedule.html"

    def get(self, request, workspace_id):
        workspace = Workspace.objects.get(id=workspace_id)
        schedule = WorkspaceSchedule.objects.get(workspace__id=workspace_id)
        form = ScheduleForm(initial={'hours': schedule.interval_hours})
        form.fields['start_datetime'].widget.js_options['defaultDate'] = schedule.start_datetime.strftime(
            '%Y-%m-%d %I:%M %p')
        context = {"schedule": "active", "workspace_id": workspace_id,
                   "workspace_name": workspace.name, "form": form,
                   "enabled": schedule.enabled, "settings_tab": "active"}
        return render(request, self.template_name, context)

    def post(self, request, workspace_id):
        schedule = WorkspaceSchedule.objects.get(workspace__id=workspace_id)
        start_datetime = request.POST.get('start_datetime')
        schedule.start_datetime = make_aware(datetime.strptime(start_datetime, '%Y-%m-%d %I:%M %p'))
        schedule.interval_hours = request.POST.get('hours')
        enabled = request.POST.get('schedule')
        schedule.enabled = False
        if enabled == 'true':
            schedule.enabled = True
        schedule.save()

        if schedule.enabled:
            schedule_sync(
                workspace_id=workspace_id,
                schedule=schedule,
                user=request.user
            )

        return HttpResponseRedirect(self.request.path_info)


class ScheduleSyncView(APIView):
    """
    Schedule task view
    """

    def post(self, request, workspace_id):
        run_sync_schedule(workspace_id, request.user)

        return Response(status=status.HTTP_200_OK)
